"""Export campaign reports to Excel and PDF."""
from pathlib import Path
from app.core.logger import get_logger

log = get_logger(__name__)


def _build_rows(campaign_id: int) -> tuple[str, list[dict]]:
    from app.core.database import get_session
    from app.models.campaign import Campaign, CampaignSend
    from app.models.contact import Contact

    with get_session() as s:
        camp = s.get(Campaign, campaign_id)
        if not camp:
            raise ValueError(f"Campaign {campaign_id} not found")
        title = f"{camp.name} — Campaign Report"
        sends = s.query(CampaignSend).filter_by(campaign_id=campaign_id).all()
        rows = []
        for send in sends:
            contact = s.get(Contact, send.contact_id)
            rows.append({
                "Email": contact.email if contact else "?",
                "First Name": (contact.first_name or "") if contact else "",
                "Last Name": (contact.last_name or "") if contact else "",
                "Company": (contact.company or "") if contact else "",
                "Status": send.status,
                "Sent At": str(send.sent_at or "")[:16],
                "Opened At": str(send.opened_at or "")[:16],
                "Message ID": send.message_id or "",
                "Error": send.error_message or "",
            })
    return title, rows


def export_excel(campaign_id: int, output_path: Path) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    title, rows = _build_rows(campaign_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    if rows:
        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="313244")
        header_font = Font(bold=True, color="89B4FA")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for r_idx, row in enumerate(rows, 3):
            fill = PatternFill("solid", fgColor="1E1E2E" if r_idx % 2 else "181825")
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
                cell.fill = fill

        # Auto-width
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    wb.save(str(output_path))
    log.info("Excel report: %s", output_path)
    return output_path


def export_pdf(campaign_id: int, output_path: Path) -> Path:
    from app.services.document.pdf_service import generate_campaign_report_pdf
    title, rows = _build_rows(campaign_id)
    return generate_campaign_report_pdf(rows, title, output_path)


def export_summary_excel(campaigns_data: list[dict], output_path: Path) -> Path:
    """Multi-campaign summary sheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ["Campaign", "Status", "Sent", "Opens", "Clicks", "Bounces", "Open Rate"]
    hfill = PatternFill("solid", fgColor="313244")
    hfont = Font(bold=True, color="89B4FA")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont

    for r, d in enumerate(campaigns_data, 2):
        sent = d.get("sent_count", 0)
        opens = d.get("open_count", 0)
        rate = f"{opens/sent*100:.1f}%" if sent else "—"
        ws.append([
            d.get("name"), d.get("status"), sent, opens,
            d.get("click_count", 0), d.get("bounce_count", 0), rate,
        ])

    wb.save(str(output_path))
    return output_path
