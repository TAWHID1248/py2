"""QThread worker for importing contacts from Excel/CSV."""
from pathlib import Path
from PySide6.QtCore import QThread, Signal
from app.core.database import get_session
from app.repositories.contact_repository import ContactRepository
from app.core.logger import get_logger

log = get_logger(__name__)


class ImportWorker(QThread):
    progress = Signal(int, int)          # current, total
    finished = Signal(int, str)          # rows_imported, list_name
    error = Signal(str)

    def __init__(self, file_path: str, list_name: str, parent=None):
        super().__init__(parent)
        self.file_path = Path(file_path)
        self.list_name = list_name

    def run(self):
        try:
            rows = self._load_rows()
            total = len(rows)
            imported = 0

            with get_session() as s:
                repo = ContactRepository(s)
                cl = repo.get_list_by_name(self.list_name)
                if not cl:
                    cl = repo.create_list(self.list_name)

                for i, row in enumerate(rows):
                    email = row.get("email", "").strip()
                    if not email or "@" not in email:
                        continue
                    if repo.is_suppressed(email):
                        continue
                    contact, _ = repo.upsert_contact(row)
                    repo.add_to_list(cl.id, contact.id)
                    imported += 1
                    self.progress.emit(i + 1, total)

            self.finished.emit(imported, self.list_name)

        except Exception as exc:
            log.exception("Import error: %s", exc)
            self.error.emit(str(exc))

    def _load_rows(self) -> list[dict]:
        suffix = self.file_path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv()
        elif suffix in (".xlsx", ".xls"):
            return self._load_excel()
        raise ValueError(f"Unsupported file type: {suffix}")

    def _load_csv(self) -> list[dict]:
        import csv
        with open(self.file_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [self._normalise(row) for row in reader]

    def _load_excel(self) -> list[dict]:
        suffix = self.file_path.suffix.lower()
        if suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(self._normalise(dict(zip(headers, row))))
            wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(str(self.file_path))
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]
            rows = []
            for r in range(1, ws.nrows):
                row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
                rows.append(self._normalise(row))
        return rows

    def _normalise(self, row: dict) -> dict:
        mapping = {
            "email": ["email", "e-mail", "email address"],
            "first_name": ["first_name", "first name", "firstname", "fname"],
            "last_name": ["last_name", "last name", "lastname", "lname", "surname"],
            "company": ["company", "organisation", "organization", "company name"],
            "phone": ["phone", "phone number", "telephone", "mobile"],
        }
        result = {}
        lower_row = {str(k).lower().strip(): v for k, v in row.items()}
        for field, aliases in mapping.items():
            for alias in aliases:
                if alias in lower_row and lower_row[alias]:
                    result[field] = str(lower_row[alias]).strip()
                    break
        return result
